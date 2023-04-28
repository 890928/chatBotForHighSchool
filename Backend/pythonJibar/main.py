from flask import Flask,request,json
from flask_restful import Api,Resource

import jieba
import jieba.analyse
import openpyxl

app = Flask(__name__)
@app.route('/GetMessage/Message',methods=['GET'])
def GetMessage():
    if request.method == 'GET':
        Message = request.args.get('message')
        reqtaxa=request.args.get('taxa')
        chi= request.args.get('chi')
        eng= request.args.get('eng')
        math= request.args.get('math')
        major1= request.args.get('major1')
        major2= request.args.get('major2')
        aptitu= request.args.get('aptitu')
        jieba.load_userdict('analyzeWord.txt')
        #停用字處理
        seg_list =jieba.lcut(Message,cut_all=False)
        file_name = 'chineseStopWord.txt'
        seg_list = remove_stop_words(file_name,seg_list)
        #組合起來進行TFIDF
        seg_list=getKeyWord(seg_list)
        #檢查是否有科系推薦
        if(checkAptituWord(seg_list)):
           return{
               'aptitu':aptituAnalyze(aptitu)
           }
        if(checkFutureWord(seg_list)):
           return{
               'future':futureAnalyze(aptitu)
           }
        #檢查是否有類群
        taxa =checkTaxaWord(seg_list)
        if taxa:
            return taxa
        #檢查分數種類
        point =checkPointWord(seg_list)
        if point:
            return point
        #檢查是否有性向相關
        if(checkInterest('interestWord.txt',seg_list)):
            return {
                'message': '請您完成性向測驗'
            }
        #檢查是否有發展相關
        
        #落點分析
        if(reqtaxa is None):
            return {
                'message': '請告訴小寶您的考試類群'
            }
        elif(chi is None):
            return {
                'message': '請告訴小寶您的國文分數'
            }
        elif(eng is None):
            return{
                'message': '請告訴小寶您的英文分數'
            }
        elif(math is None):
            return{
                'message': '請告訴小寶您的數學分數'
            }
        elif(major1 is None):
            return{
                'message': '請告訴小寶您的專一分數'
            }
        elif(major2 is None):
            return{
                'message': '請告訴小寶您的專二分數'
            }
        school = checkSchoolWord(seg_list)
        department =checkDepartmentWord(seg_list)
        if(school is not None and department is not None):
            if(reqtaxa is None or chi is None or eng is None or math is None or major1 is None or major2 is None):
                return {
                    'message': '分數或類群資料尚未齊全喔!',
                }
            else :
                pointData =pointAnalyze(reqtaxa,school,department)
                totalPoint =totalPointAnalyze(pointData,chi,eng,math,major1,major2)
                resultData = tidyPointAnalyze(pointData,totalPoint)
                if not pointData:
                    return {
                        'message':'當前類群無此校系資料'
                    }
                else:
                    return {
                        'resultSchool': resultData
                    }
        return {
            'message': '請告訴小寶您想要就讀的學校及科系'
        }
    
# 落點分析
def pointAnalyze(reqtaxa,school,department):
    wb = openpyxl.load_workbook('pointAnalyze.xlsx')  # 設定 data_only=True 只讀取計算後的數值
    s1 = wb[reqtaxa]
    arr = []                      # 第一層串列
    for row in s1:
        arr2 = []                 # 第二層串列
        schoolCheck='N'
        departmentCheck='N'
        for column in row:
            if(column.value == school):
                schoolCheck='Y'
            if(column.value == department):
                departmentCheck='Y'
            arr2.append(column.value)  # 寫入內容
        if(schoolCheck == 'Y' and departmentCheck == 'Y'):
            arr.append(arr2)
    return arr
# 計算加權平均分數
def totalPointAnalyze(arr,chi,eng,math,major1,major2):
    for row in arr:
        tidyMath=row[2]
    tidyMath = tidyMath.split(' + ')
    for row in tidyMath:
        tmpRow = row.split("*")
        if(tmpRow[0] == '國文'):
            chi = float(chi)*float(tmpRow[1])
        elif (tmpRow[0] == '英文'):
            eng = float(eng)*float(tmpRow[1])
        elif (tmpRow[0] == '數學'):
            math = float(math)*float(tmpRow[1])
        elif (tmpRow[0] == '專業(一)'):
            major1 = float(major1)*float(tmpRow[1])
        elif (tmpRow[0] == '專業(二)'):
            major2 = float(major2)*float(tmpRow[1])
    totalPoint = chi+eng+math+major1+major2
    return totalPoint
# 整理回傳學校資料
def tidyPointAnalyze(arr,totalPoint):
    for row in arr:
        arr=row
    arr[2]='歷年錄取分數:'+str(arr[3])
    arr[3]='您的加權總分:'+str(totalPoint)
    return arr
        
# 檢查是否有類群
def checkTaxaWord(seg_list):
    with open('taxaWord.txt','r',encoding="utf-8") as f:
        taxa_words = f.readlines()
    taxa_words = [taxa_word.rstrip() for taxa_word in taxa_words]
    new_list = []
    for seg in seg_list:
        if seg in taxa_words:
            return {'storgeName':'keepTaxa','message':seg}
    return {}

# 檢查是否有分數
def checkPointWord(seg_list):
    with open('pointWord.txt','r',encoding="utf-8") as f:
        taxa_words = f.readlines()
    taxa_words = [taxa_word.rstrip() for taxa_word in taxa_words]
    new_list = []
    for seg in seg_list:
        if seg in taxa_words:
            for seg_point in seg_list:
                if seg_point.isdigit():
                    return {'storgeName':seg,'message':seg_point}
    return []

# 檢查是否有科系
def checkDepartmentWord(seg_list):
    with open('departmentWord.txt','r',encoding="utf-8") as f:
        department_words = f.readlines()
    department_words = [department_word.rstrip() for department_word in department_words]
    new_list = []
    for seg in seg_list:
        if seg in department_words:
            return seg
    return None

# 檢查是否有學校
def checkSchoolWord(seg_list):
    with open('schoolWord.txt','r',encoding="utf-8") as f:
        school_words = f.readlines()
    school_words = [school_word.rstrip() for school_word in school_words]
    new_list = []
    for seg in seg_list:
        if seg in school_words:
            return seg
    return None
# 檢查科系推薦
def checkAptituWord(seg_list):
    with open('aptituWord.txt','r',encoding="utf-8") as f:
        aptitu_words = f.readlines()
    aptitu_words = [aptitu_word.rstrip() for aptitu_word in aptitu_words]
    new_list = []
    for seg in seg_list:
        if seg in aptitu_words:
            return True
    return False
# 科系推薦資料
def aptituAnalyze(aptitu):
    wb = openpyxl.load_workbook('aptituAnalyze.xlsx')  # 設定 data_only=True 只讀取計算後的數值
    s1 = wb[aptitu]
    arr = []                      # 第一層串列
    for row in s1:
        arr2 = []                 # 第二層串列
        for column in row:
            arr2.append(column.value)  # 寫入內容
        arr.append(arr2)
    return arr
# 檢查未來出路
def checkFutureWord(seg_list):
    with open('futureWord.txt','r',encoding="utf-8") as f:
        future_words = f.readlines()
    future_words = [future_word.rstrip() for future_word in future_words]
    new_list = []
    for seg in seg_list:
        if seg in future_words:
            return True
    return False
# 科系推薦資料
def futureAnalyze(aptitu):
    wb = openpyxl.load_workbook('futureAnalyze.xlsx')  # 設定 data_only=True 只讀取計算後的數值
    s1 = wb[aptitu]
    arr = []                      # 第一層串列
    for row in s1:
        arr2 = []                 # 第二層串列
        for column in row:
            arr2.append(column.value)  # 寫入內容
        arr.append(arr2)
    return arr
# 檢查是否為性向字詞
def checkInterest(file_name,seg_list):
    with open(file_name,'r',encoding="utf-8") as f:
        interest_words = f.readlines()
    interest_words = [interest_word.rstrip() for interest_word in interest_words]
    new_list = []
    for seg in seg_list:
        if seg in interest_words:
            return True
    return False

# 移除停留詞
def remove_stop_words(file_name,seg_list):
  with open(file_name,'r',encoding="utf-8") as f:
    stop_words = f.readlines()
  stop_words = [stop_word.rstrip() for stop_word in stop_words]
  new_list = []
  for seg in seg_list:
    if seg not in stop_words:
      new_list.append(seg) #若在for loop裡用remove的話則會改變總長度
  return new_list

#TFIDF處理
def getKeyWord(seg_list):
    text=''
    for seg in seg_list:
        text+=seg
    text =jieba.analyse.extract_tags(text, topK=5)
    return text




if __name__ == "__main__":
    # 0.0.0.0任何機器地址都訪問
    app.run(host='192.168.172.96',port='3000')  # 運行程序